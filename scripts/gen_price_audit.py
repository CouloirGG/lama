"""Generate LAMA Price Audit spreadsheet for cross-checking app vs in-game prices."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Price Audit"

# Colors
DARK_BG = "FF1c1814"
GOLD = "FFc4a456"
HEADER_BG = "FF2a2318"
CATEGORY_BG = "FF12100c"
WHITE = "FFd4c9a8"
MUTED = "FF8c7a5c"
RED = "FFa83232"
GREEN = "FF4a7c59"
BORDER_COLOR = "3a3128"

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

# Column widths
for col, w in {"A": 22, "B": 30, "C": 16, "D": 16, "E": 16, "F": 12, "G": 30}.items():
    ws.column_dimensions[col].width = w

# Styles
title_font = Font(name="Segoe UI", size=14, bold=True, color=GOLD)
header_font = Font(name="Segoe UI", size=10, bold=True, color=GOLD)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
cat_font = Font(name="Segoe UI", size=11, bold=True, color=GOLD)
cat_fill = PatternFill(start_color=CATEGORY_BG, end_color=CATEGORY_BG, fill_type="solid")
data_font = Font(name="Consolas", size=10, color=WHITE)
muted_font = Font(name="Consolas", size=10, color=MUTED)
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

# Fill background
for row in range(1, 200):
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = dark_fill
        cell.font = data_font

# Title
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "LAMA Price Audit \u2014 Cross-check vs In-Game Trade"
c.font = title_font
c.fill = dark_fill
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

# Instructions
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = (
    "Fill in App Price (from overlay/dashboard) and In-Game Price "
    "(from trade site). Disparity auto-calculates. Add notes for context."
)
c.font = Font(name="Segoe UI", size=9, italic=True, color=MUTED)
c.fill = dark_fill
ws.row_dimensions[2].height = 20

# Date/league row
ws.merge_cells("A3:G3")
c = ws["A3"]
c.value = "Date: ____________    League: ____________    Tester: ____________"
c.font = Font(name="Segoe UI", size=10, color=MUTED)
c.fill = dark_fill
ws.row_dimensions[3].height = 22

# Headers
headers = [
    "Category", "Item Name", "App Price (exalted)",
    "In-Game Price (exalted)", "Disparity (exalted)", "Disparity %", "Notes",
]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_idx)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[5].height = 24

# Currency categories with representative items
categories = {
    "Currency": [
        "Divine Orb", "Exalted Orb", "Chaos Orb", "Vaal Orb",
        "Orb of Alchemy", "Orb of Annulment", "Fracturing Orb",
        "Mirror of Kalandra", "Regal Orb", "Orb of Chance",
    ],
    "Fragments": [
        "Simulacrum Splinter", "Breach Splinter", "Tablet Fragment",
        "Precursor Tablet", "Maven Splinter",
    ],
    "Essences": [
        "Greater Essence of the Body", "Greater Essence of Haste",
        "Greater Essence of Sorcery", "Essence of the Body",
        "Essence of Flames", "Essence of Ice",
    ],
    "Runes": [
        "Iron Rune", "Desert Rune", "Body Rune",
        "Mind Rune", "Storm Rune", "Glacial Rune",
    ],
    "Expedition": [
        "Exotic Coinage", "Astragali", "Scrap Metal",
        "Burial Medallion",
    ],
    "Soul Cores": [
        "Soul Core of Tacati", "Soul Core of Anapotia",
        "Soul Core of Topotante", "Soul Core of Puhuarte",
        "Soul Core of Zalatl", "Soul Core of Opiloti",
    ],
    "Uncut Gems": [
        "Uncut Skill Gem (lvl 16)", "Uncut Skill Gem (lvl 19)",
        "Uncut Skill Gem (lvl 20)", "Uncut Spirit Gem (lvl 16)",
        "Uncut Spirit Gem (lvl 20)", "Uncut Support Gem (lvl 16)",
    ],
    "Ultimatum": [
        "Inscribed Ultimatum", "Catalyst of Corruption",
        "Trial Coins",
    ],
    "Catalysts": [
        "Tempering Catalyst", "Prismatic Catalyst",
        "Fertile Catalyst", "Intrinsic Catalyst",
        "Turbulent Catalyst", "Abrasive Catalyst",
    ],
    "Delirium": [
        "Distilled Ire", "Distilled Paranoia",
        "Distilled Despair", "Distilled Guilt",
        "Distilled Greed", "Distilled Disgust",
    ],
    "Ritual": [
        "Blood-filled Vessel", "Ritual Vessel",
        "Ritual Splinter",
    ],
}

row = 6
for category, items in categories.items():
    # Category header row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1)
    cell.value = category.upper()
    cell.font = cat_font
    cell.fill = cat_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    for item in items:
        ws.cell(row=row, column=1, value=category).font = muted_font
        ws.cell(row=row, column=1).fill = dark_fill
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=item).font = data_font
        ws.cell(row=row, column=2).fill = dark_fill
        ws.cell(row=row, column=2).border = thin_border

        # App price (user fills in)
        c = ws.cell(row=row, column=3)
        c.font = data_font
        c.fill = dark_fill
        c.border = thin_border
        c.number_format = "#,##0.0"
        c.alignment = Alignment(horizontal="center")

        # In-game price (user fills in)
        c = ws.cell(row=row, column=4)
        c.font = data_font
        c.fill = dark_fill
        c.border = thin_border
        c.number_format = "#,##0.0"
        c.alignment = Alignment(horizontal="center")

        # Disparity formula: In-Game - App
        c = ws.cell(row=row, column=5)
        c.value = f'=IF(AND(C{row}<>"",D{row}<>""),D{row}-C{row},"")'
        c.font = data_font
        c.fill = dark_fill
        c.border = thin_border
        c.number_format = "+#,##0.0;-#,##0.0;0"
        c.alignment = Alignment(horizontal="center")

        # Disparity % formula
        c = ws.cell(row=row, column=6)
        c.value = f'=IF(AND(C{row}<>"",D{row}<>"",C{row}<>0),(D{row}-C{row})/C{row},"")'
        c.font = data_font
        c.fill = dark_fill
        c.border = thin_border
        c.number_format = "+0.0%;-0.0%;0%"
        c.alignment = Alignment(horizontal="center")

        # Notes
        c = ws.cell(row=row, column=7)
        c.font = muted_font
        c.fill = dark_fill
        c.border = thin_border

        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # spacer

# Conditional formatting — highlight >10% disparity in red, <-10% in green
ws.conditional_formatting.add(
    f"F7:F{row}",
    CellIsRule(operator="greaterThan", formula=["0.1"],
              font=Font(color=RED, bold=True)),
)
ws.conditional_formatting.add(
    f"F7:F{row}",
    CellIsRule(operator="lessThan", formula=["-0.1"],
              font=Font(color=GREEN, bold=True)),
)

# Freeze header row
ws.freeze_panes = "A6"

out = str(__import__("pathlib").Path(__file__).resolve().parent.parent / "LAMA_Price_Audit.xlsx")
wb.save(out)
print(f"Created: {out}")
